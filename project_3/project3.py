import argparse
import pymongo
import subprocess
import shlex
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import os

#set up argparse
parser = argparse.ArgumentParser()
parser.add_argument('--files', dest='bl_flames', nargs='*')
parser.add_argument('--xytech', dest='xytech')
# 2. Add/Run new argparse command --process <video file> to process provided video file
parser.add_argument('--process', dest='video')
parser.add_argument('--verbose', action='store_true')
parser.add_argument('--output', dest='output')

args = parser.parse_args()

#set up code for database calls
myclient = pymongo.MongoClient('mongodb://localhost:27017')
mydb = myclient['mydatabase']
col1 = mydb['collection1']
col2 = mydb['collection2']

#function to convert timecode to frames
#revised from lesson 7 to take in fps argument
def frames_to_TC (frames, fps):
    h = frames / (fps * 60 * 60)
    m = (frames / (fps*60)) % 60 
    s = (frames % (fps*60))/fps
    f = frames % fps
    return ( "%02d:%02d:%02d:%02d" % ( h, m, s, f))

#function to convert timecode, retrieved from ffmpeg command, to frames
#takes in String timecode and Int fps for arguments
def TC_to_frames(timecode, fps):
    h, m, s = timecode.split(':')
    if '.' in s:
        s = int(s.split('.')[0])
        hundreths = int(timecode.split('.')[-1])
        total_secs = (int(h) * 3600) + (int(m) * 60) + s + (hundreths/100)
    else:
        total_secs = (int(h) * 3600) + (int(m) * 60) + int(s)
    frames = math.floor(total_secs * fps)
    return frames


video = args.video
if video.startswith('.\\'):
    video = video.removeprefix('.\\')

#full line command for video info
vid_info_command = f"ffmpeg -i {video} -hide_banner"

process = subprocess.Popen(
    shlex.split(vid_info_command),
    stdout=subprocess.PIPE,
    stderr=subprocess.STDOUT,
    shell=True
)

#Preparation for 3.

#Get video duration
info = process.stdout.readlines() 
vid_duration = info[6].decode().strip()
vid_duration_list = vid_duration.split(',')
vid_duration = vid_duration_list[0].removeprefix('Duration: ')

#Get frames per second (fps) from video
frames = info[7].decode().strip()
frames = frames.split(",")
fps = int(frames[6].replace(" fps", ""))

#Convert timecode retrieved from video to frames
vid_duration_to_frames = TC_to_frames(vid_duration, fps)

within_vid = []
# 3. Call the populated database from proj2, find all ranges only that fall in the length of video from 1
all_docs = col2.find({})

for doc in all_docs:
    #separate location and frame(s), put in list
    for loc_and_frames in doc['Frames to fix']:
        loc_frame_split = loc_and_frames.split(" ")
        #if statement to check for frame ranges only
        if '-' in loc_frame_split[1]:
            range_split = loc_frame_split[1].split('-')
            l_range = int(range_split[0])
            r_range = int(range_split[1])
            #if the left or right frames of the range go past the length of video, continue to next iteration
            if l_range > vid_duration_to_frames or r_range > vid_duration_to_frames:
                continue
            else:
                # 4. Convert marks into timecodes
                l_range_TC = frames_to_TC(l_range, fps)
                r_range_TC = frames_to_TC(r_range, fps)
                range_to_TC = l_range_TC+'-'+r_range_TC
                within_vid.append([loc_frame_split[0],loc_frame_split[1],range_to_TC])

        else:
            continue

if args.output:
    #5. New argparse --output parameter to run with --process
    #create the .xlsx file/workbook

    for row in within_vid:
        frame_range = row[1]
        fr_separate = frame_range.split("-")
        mid_frame = math.ceil(int((int(fr_separate[0])+int(fr_separate[1]))/2))
        #convert middle frame to timecode
        mid_frame_TC = frames_to_TC(mid_frame, fps)
        #convert middle frame to total seconds
        #use this to capture point in time to grab thumbnail
        mid_frame_to_secs = mid_frame*(1/fps)
        #set width and output file location/name
        width = '96x74'
        loc = row[0].removeprefix('/').replace('/','-')
        output = f'images/{loc}_{str(mid_frame)}.png'
        #if file exists, don't run the command and continue to next iteration
        if os.path.isfile(output):
            #append string image_file to within_vid list to be used later when creating excel file
            #within_vid should now contain nested lists holding these elements:
            #[<location>, <Ranges within vid>, <Timecode HH:MM:SS:FF>, <image file or path/to/image file>]
            row.append(output)
            print(f'File exists. Location/name: {output}')
            continue
        else:
            row.append(output)
            #set up command to get single frame for thumbnail
            #pulls single frame from given timestamp (in seconds or timecode format), saves file to a given file name/folder location
            #also formats the image to have 96x74 resolution size
            get_thumb = f'ffmpeg -i {args.video} -frames:v 1 -s 96x74 -ss {mid_frame_to_secs} {output}'
            #using subprocess.run to run command asynchronously, making sure all thumbnails are created before including in .xlsx file
            result = subprocess.run(get_thumb, capture_output=True, text=True)
            #check to see if command ran with any errors or not
            if result.returncode == 0:
                img_loc_name = output.split('/')
                file_name = img_loc_name[-1]
                img_loc_name.pop()
                #console output to show name of image file and/or current directory to image file
                print(f'New image [{file_name}] was generated at sub-folder location: {img_loc_name}')
            else:
                print('command failed')

    print('Thumbnail creation completed.')

    #start making excel file
    wb = Workbook()
    ws = wb.active
    #set sheet name and column names, format as previous csv output, except include thumbnail
    ws.title = 'Mark-ups'
    ws['A1'] = 'Location'
    ws['B1'] = 'Ranges'
    ws['C1'] = 'Timecode'
    ws['D1'] = 'Thumbnail'

    #format cell dimensions
    ws.sheet_format.defaultRowHeight = 50
    ws.sheet_format.defaultColWidth = 50
    for row_index, lst in enumerate(within_vid, start=2):
        for col_index, col_value in enumerate(lst, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            #following code inserts thumbnail at current row, 4th column
            if col_index == 4:
                #col_value should be image_file, or path/to/image_file
                #load image/thumbnail using Image()
                img = Image(col_value)
                ws.column_dimensions[cell.column_letter].width = 50
                ws.row_dimensions[cell.row].height = 70
                #insert image in cell
                ws.add_image(img, cell.coordinate)
            else:
                #inserts <location>, <frame ranges>, and/or <timecode>
                cell.value = col_value

    #create the excel workbook/ .xlxs file
    file_name = f'{args.output}.xlsx'
    wb.save(file_name)